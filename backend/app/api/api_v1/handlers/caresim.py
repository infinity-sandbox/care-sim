from datetime import datetime
from fastapi import Depends, FastAPI, File, HTTPException, Header, Request, UploadFile, status
from fastapi.responses import FileResponse, JSONResponse
from grpc import Status
import jwt
from pydantic import BaseModel
from typing import List, Optional
import openai
import os
import json
from fastapi import APIRouter
from app.core.config import logger_settings
logger = logger_settings.get_logger(__name__)
import aiomysql
import asyncio
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from fastapi.responses import StreamingResponse
from app.services.auth_service import AuthDatabaseService
# Add to your FastAPI router
import math
from fastapi import FastAPI, HTTPException
from pydantic import BaseModel
from typing import List, Optional
import smtplib
from email.mime.text import MIMEText
from email.mime.multipart import MIMEMultipart
from reportlab.lib.pagesizes import letter
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
import uuid
from utils.vapor.engine.scheduler import start_scheduler

insight_router = APIRouter()

# Define Pydantic models for input data
class RevenueSource(BaseModel):
    id: Optional[str] = None
    sourceName: str
    monthlyAmount: float
    tag: str

class EmployeeExpense(BaseModel):
    id: Optional[str] = None
    expenseName: str
    monthlyAmount: float
    type: str
    hoursPerMonth: Optional[float] = None
    
class FacilityExpense(BaseModel):
    id: Optional[str] = None
    expenseName: str
    monthlyAmount: float
    type: str

class Classroom(BaseModel):
    id: Optional[str] = None
    name: str
    capacity: int
    ratio: float
    avgStudents: int

class OperatingDetail(BaseModel):
    operatingHours: float
    operatingDays: float

    
class Administrative(BaseModel):
    id: Optional[str] = None
    expenseName: str
    monthlyAmount: float
    type: str

class Supplies(BaseModel):
    id: Optional[str] = None
    expenseName: str
    monthlyAmount: float
    type: str
    
class Goals(BaseModel):
    id: Optional[str] = None
    goal: str
    targetPercentage: float

class DaycareInput(BaseModel):
    businessName: Optional[str] = "Unamed Daycare"
    revenueSources: List[RevenueSource] = []
    employees: List[EmployeeExpense] = []
    facilities: List[FacilityExpense] = []
    administrative: List[FacilityExpense] = []
    supplies: List[FacilityExpense] = []
    classrooms: List[Classroom] = []
    operatingHours: Optional[float] = 0.0  # Default operating hours
    operatingDays: Optional[float] = 0.0  # Default operating days
    # operatingDetails: List[OperatingDetail] = []
    goals: List[Goals] = []

    
# Function to generate insights using OpenAI
async def generate_daycare_insights(data: DaycareInput) -> dict:
    # Prepare the prompt
    prompt = f"""
    You are a financial and operations analyst for daycare centers. Analyze the following daycare center data and provide insights in the specified JSON format:

    Daycare Name: {data.businessName}

    REVENUE SOURCES:
    {json.dumps([src.dict() for src in data.revenueSources], indent=2)}

    EXPENSES:
    - Employees: {json.dumps([emp.dict() for emp in data.employees], indent=4)}
    - Facilities: {json.dumps([fac.dict() for fac in data.facilities], indent=4)}
    - Administrative: {json.dumps([adm.dict() for adm in data.administrative], indent=4)}
    - Supplies: {json.dumps([sup.dict() for sup in data.supplies], indent=4)}

    CLASSROOM DETAILS:
    {json.dumps([cls.dict() for cls in data.classrooms], indent=2)}

    OPERATING DETAILS:
    {json.dumps({"operatingHours": data.operatingHours, "operatingDays": data.operatingDays}, indent=2)}

    BUSINESS GOALS:
    {json.dumps([goal.dict() for goal in data.goals], indent=2)}

    Provide your analysis in the following JSON format:
    (Return the following format as valid JSON **only**, without any extra text or explanation)

    {{
      "net_monthly_income": {{
        "value": number,
        "currency": "USD",
        "calculation": "string explaining the calculation",
        "note": "brief interpretation"
      }},
      "break_even_enrollment": {{
        "value": number,
        "unit": "students",
        "calculation": "string explaining the calculation",
        "note": "brief interpretation"
      }},
      "largest_expense": {{
        "category": "string",
        "percentage_of_total_expenses": number,
        "calculation": "string explaining the calculation"
      }},
      "capacity_utilization": {{
        "value": number,
        "unit": "percent",
        "calculation": "string explaining the calculation",
        "note": "brief interpretation"
      }},
      "executive_summary": {{
        "financial_overview": "string",
        "profitability_status": "string",
        "enrollment_status": "string",
        "recommendations": ["string", "string", "string", "string"]
      }}
    }}

    Key calculations to include:
    1. Net Monthly Income = Total Revenue - Total Expenses
    2. Break-Even Enrollment = Total Fixed Costs / Average Revenue per Student
    3. Largest Expense Category = Expense category with highest percentage of total expenses
    4. Capacity Utilization = (Total Enrolled Students / Total Capacity) * 100

    Make all calculations based on monthly figures. Provide realistic, actionable insights.
    """
    
    try:
        client = openai.AsyncOpenAI(api_key=logger_settings.OPENAI_API_KEY)
        model = "gpt-4-turbo"  # or "gpt-4-turbo" if available
        messages=[
                {"role": "system", "content": "You are a financial analyst specializing in daycare center operations."},
                {"role": "user", "content": prompt}
            ]
        print(f"Generating insights for {data.businessName}...")
    
        response = await client.chat.completions.create(
            model=model,
            messages=messages,
            max_tokens=1500,
            temperature=0.8,
            response_format={"type": "json_object"}
        )
        
        # Extract the actual content from the response
        insights_content = response.choices[0].message.content
        
        # Parse the JSON content
        insights_data = json.loads(insights_content)
        print(f"Insights content: \n\n{insights_data}")
        return insights_data
        
    except json.JSONDecodeError as e:
        print(f"JSON decode error: {str(e)}")
        print(f"Problematic content: {insights_content}")
        raise HTTPException(
            status_code=500, 
            detail=f"Failed to parse OpenAI response: {str(e)}"
        )
    except Exception as e:
        raise HTTPException(
            status_code=500, 
            detail=f"OpenAI API error: {str(e)}"
        )
    
async def verify_token(authorization: str = Header(...)):
    """Dependency to verify and extract token"""
    if not authorization.startswith("Bearer "):
        raise HTTPException(
            status_code=401,
            detail="Invalid authorization header"
        )
    
    token = authorization.split(" ")[1]
    
    try:
        payload = jwt.decode(token, logger_settings.JWT_SECRET_KEY, algorithms=[logger_settings.ALGORITHM])
        return payload
    except jwt.ExpiredSignatureError:
        raise HTTPException(
            status_code=401,
            detail="Token expired"
        )
    except jwt.InvalidTokenError:
        raise HTTPException(
            status_code=401,
            detail="Invalid token"
        )
        
        
# Endpoint to generate insights
@insight_router.post("/generate-insights")
async def get_insights(input_data: DaycareInput,
            request: Request,
            token_payload: dict = Depends(verify_token),
            db: aiomysql.Connection = Depends(AuthDatabaseService.get_db)):
    
    try:
        await start_scheduler()
        async with db.cursor(aiomysql.DictCursor) as cursor:
            # Check if the user exists based on email
            await cursor.execute(
                "SELECT * FROM auth_users WHERE id = %s",
                (token_payload.get("sub"),)
            )
        user = await cursor.fetchone()  # Fetch the first matching user
        logger.info(f"User: {user['email']}")
        if not user:
            raise HTTPException(
                status_code=401,
                detail="Invalid token: user ID not found"
            )
        # Example usage
        data = {
            "net_monthly_income": {
                "value": 3000,
                "currency": "USD",
                "calculation": "Total Revenue (45.0) - Total Expenses (3785.0)",
                "note": "The daycare center is currently operating at a loss, indicating a need to either increase revenue or reduce expenses."
            },
            "break_even_enrollment": {
                "value": 1,
                "unit": "students",
                "calculation": "Total Fixed Costs (3785.0) / Average Revenue per Student (45.0)",
                "note": "To cover its expenses, the daycare would need to enroll approximately 84 students, which exceeds its current capacity."
            },
            "largest_expense": {
                "category": "Supplies",
                "percentage_of_total_expenses": 30.7,
                "calculation": "(Food Expense 3434.0 / Total Expenses 3785.0) * 100"
            },
            "capacity_utilization": {
                "value": 40,
                "unit": "percent",
                "calculation": "(Total Enrolled Students 34 / Total Capacity 34) * 100",
                "note": "The daycare is fully utilizing its capacity, which is a positive indicator for enrollment management."
            },
            "executive_summary": {
                "financial_overview": "The daycare center is currently facing financial challenges with a significant monthly loss.",
                "profitability_status": "Not profitable, operating at a loss.",
                "enrollment_status": "Fully enrolled at current capacity.",
                "recommendations": [
                    "Consider increasing revenue through additional services or programs.",
                    "Review and reduce supply costs to improve overall profitability.",
                    "Review and reduce supply costs to improve overall profitability.",
                    "Review and reduce supply costs to improve overall profitability.",
                    "Consider increasing revenue through additional services or programs."
                ]
            }
        }
        insight = await generate_daycare_insights(input_data)
        conn = await AuthDatabaseService.connection()
        try:
            async with conn.cursor() as cursor:
                # Convert Pydantic model to dict, then to JSON string
                data_json = json.dumps(insight)
                
                await cursor.execute(
                    "INSERT INTO insights_data (user_email, data) VALUES (%s, %s)",
                    (user["email"], data_json)
                )
                await conn.commit()
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error saving insight: {e}")
        finally:
            conn.close()
        # In a real implementation, you would save to a database
        logger.info("Insights saved successfully")
        return insight
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))


# Endpoint to save inputs (optional)
@insight_router.post("/save-inputs")
async def save_inputs(input_data: DaycareInput,
                      request: Request,
                      token_payload: dict = Depends(verify_token),
                      db: aiomysql.Connection = Depends(AuthDatabaseService.get_db)
                    ):
    """
    Save the entire DaycareInput payload as JSON into the input_data table
    """
    # Token payload contains decoded JWT claims
    async with db.cursor(aiomysql.DictCursor) as cursor:
        # Check if the user exists based on email
        await cursor.execute(
            "SELECT * FROM auth_users WHERE id = %s",
            (token_payload.get("sub"),)
        )
    user = await cursor.fetchone()  # Fetch the first matching user
    logger.info(f"User: {user['email']}")
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Invalid token: user ID not found"
        )
    conn = await AuthDatabaseService.connection()
    try:
        async with conn.cursor() as cursor:
            # Convert Pydantic model to dict, then to JSON string
            data_json = json.dumps(input_data.dict())
            
            # Insert user email and input JSON into the table
            await cursor.execute(
                "INSERT INTO input_data (user_email, data) VALUES (%s, %s)",
                (user["email"], data_json)
            )
            await conn.commit()
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error saving input: {e}")
    finally:
        conn.close()
    # In a real implementation, you would save to a database
    return {"status": "success", "message": "Inputs saved successfully"}


@insight_router.get("/fetch-inputs")
async def fetch_inputs(token_payload: dict = Depends(verify_token),
                       db: aiomysql.Connection = Depends(AuthDatabaseService.get_db)):
    pass


@insight_router.get("/recommendations")
async def recommendation(token_payload: dict = Depends(verify_token),
                       db: aiomysql.Connection = Depends(AuthDatabaseService.get_db)):
    # Token payload contains decoded JWT claims
    async with db.cursor(aiomysql.DictCursor) as cursor:
        # Check if the user exists based on email
        await cursor.execute(
            "SELECT * FROM auth_users WHERE id = %s",
            (token_payload.get("sub"),)
        )
    user = await cursor.fetchone()  # Fetch the first matching user
    logger.info(f"User: {user['email']}")
    if not user:
        raise HTTPException(
            status_code=401,
            detail="Invalid token: user ID not found"
        )
    
    async with db.cursor(aiomysql.DictCursor) as cursor:
        # Fetch the latest input_data row for the user
        await cursor.execute(
            """
            SELECT id, user_email, data, created_at 
            FROM insights_data 
            WHERE user_email = %s 
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (user['email'],)
        )
        row = await cursor.fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="No insights found for this user.")
        try:
            insights = json.loads(row['data'])
            recommendations = insights.get('executive_summary', {}).get('recommendations', [])
        except Exception:
            recommendations = []

        return {"recommendations": recommendations}
                
        
@insight_router.get("/generate-excel-report")
async def generate_excel_report(
                            token_payload: dict = Depends(verify_token),
                            db: aiomysql.Connection = Depends(AuthDatabaseService.get_db)
                        ):
    try:
        # Token payload contains decoded JWT claims
        async with db.cursor(aiomysql.DictCursor) as cursor:
            # Check if the user exists based on email
            await cursor.execute(
                """
                SELECT * 
                FROM auth_users WHERE id = %s
                """,
                (token_payload.get("sub"),)
            )
        user = await cursor.fetchone()  # Fetch the first matching user
        logger.info(f"User: {user['email']}")
        if not user:
            raise HTTPException(
                status_code=401,
                detail="Invalid token: user ID not found"
            )
        conn = await AuthDatabaseService.connection()
        try:
            async with conn.cursor(aiomysql.DictCursor) as cursor:
                # Convert Pydantic model to dict, then to JSON string
                # Insert user email and input JSON into the table
                await cursor.execute(
                    """
                    SELECT id, user_email, data, created_at
                    FROM input_data 
                    WHERE user_email = %s
                    ORDER BY created_at DESC
                    LIMIT 1
                    """,
                    (user["email"],)
                )
            row = await cursor.fetchone()  # Fetch the first matching user
            if not row:
                raise HTTPException(status_code=404, detail="No inputs found for this user.")
            
            form_data = json.loads(row['data'])
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error on input data: {e}")
        finally:
            conn.close()

        # Create a new workbook
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            default_sheet = wb['Sheet']
            wb.remove(default_sheet)
        
        # Create main title sheet
        ws_main = wb.create_sheet("Budget Summary")
        
        # Define styles
        green_fill = PatternFill(start_color="92D050", end_color="92D050", fill_type="solid")
        header_fill = PatternFill(start_color="DDEBF7", end_color="DDEBF7", fill_type="solid")
        subheader_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        bold_font = Font(bold=True)
        center_alignment = Alignment(horizontal="center", vertical="center")
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        # Add main title
        ws_main.merge_cells('A1:E1')
        title_cell = ws_main['A1']
        title_cell.value = f"Annual Budget Planner ({datetime.now().year})"
        title_cell.fill = green_fill
        title_cell.font = Font(bold=True, size=16, color="FFFFFF")
        title_cell.alignment = center_alignment
        
        # Add business information
        ws_main['A3'] = "Business Information"
        ws_main['A3'].font = bold_font
        ws_main['A3'].fill = header_fill
        
        ws_main['A4'] = "Business Name"
        ws_main['B4'] = form_data.get("businessName", "")
        
        # Add revenue sources
        row = 6
        ws_main[f'A{row}'] = "Revenue Sources"
        ws_main[f'A{row}'].font = bold_font
        ws_main[f'A{row}'].fill = header_fill
        row += 1
        
        # Revenue headers
        headers = ["Source Name", "Monthly Amount", "Tag/Note", "Yearly Amount"]
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=row, column=col)
            cell.value = header
            cell.font = bold_font
            cell.fill = subheader_fill
            cell.border = border
        
        row += 1
        total_revenue = 0
        
        # Revenue data
        for revenue in form_data.get("revenueSources", []):
            ws_main[f'A{row}'] = revenue.get("sourceName", "")
            ws_main[f'B{row}'] = revenue.get("monthlyAmount", 0)
            ws_main[f'C{row}'] = revenue.get("tag", "")
            yearly_amount = revenue.get("monthlyAmount", 0) * 12
            ws_main[f'D{row}'] = yearly_amount
            total_revenue += yearly_amount
            
            # Apply borders
            for col in range(1, 5):
                ws_main.cell(row=row, column=col).border = border
            
            row += 1
        
        # Add revenue total
        ws_main[f'A{row}'] = "Total Revenue"
        ws_main[f'A{row}'].font = bold_font
        ws_main[f'D{row}'] = total_revenue
        ws_main[f'D{row}'].font = bold_font
        for col in range(1, 5):
            ws_main.cell(row=row, column=col).border = border
        
        # Add expenses sections
        expense_categories = [
            ("employees", "Employees"),
            ("facilities", "Facilities"), 
            ("administrative", "Administrative"),
            ("supplies", "Supplies")
        ]
        
        for category_key, category_name in expense_categories:
            row += 2
            ws_main[f'A{row}'] = f"{category_name} Expenses"
            ws_main[f'A{row}'].font = bold_font
            ws_main[f'A{row}'].fill = header_fill
            row += 1
            
            # Expense headers
            headers = ["Expense Name", "Monthly Amount", "Type", "Yearly Amount"]
            if category_key == "employees":
                headers.insert(2, "Hours/Month")
            
            for col, header in enumerate(headers, 1):
                cell = ws_main.cell(row=row, column=col)
                cell.value = header
                cell.font = bold_font
                cell.fill = subheader_fill
                cell.border = border
            
            row += 1
            category_total = 0
            
            # Expense data
            for expense in form_data.get(category_key, []):
                ws_main[f'A{row}'] = expense.get("expenseName", "")
                ws_main[f'B{row}'] = expense.get("monthlyAmount", 0)
                
                if category_key == "employees":
                    ws_main[f'C{row}'] = expense.get("hoursPerMonth", "")
                    ws_main[f'D{row}'] = expense.get("type", "")
                    yearly_amount = expense.get("monthlyAmount", 0) * 12
                    ws_main[f'E{row}'] = yearly_amount
                    category_total += yearly_amount
                else:
                    ws_main[f'C{row}'] = expense.get("type", "")
                    yearly_amount = expense.get("monthlyAmount", 0) * 12
                    ws_main[f'D{row}'] = yearly_amount
                    category_total += yearly_amount
                
                # Apply borders
                col_count = 5 if category_key == "employees" else 4
                for col in range(1, col_count + 1):
                    ws_main.cell(row=row, column=col).border = border
                
                row += 1
            
            # Add category total
            total_col = 5 if category_key == "employees" else 4
            ws_main[f'A{row}'] = f"Total {category_name} Expenses"
            ws_main[f'A{row}'].font = bold_font
            ws_main.cell(row=row, column=total_col).value = category_total
            ws_main.cell(row=row, column=total_col).font = bold_font
            
            for col in range(1, total_col + 1):
                ws_main.cell(row=row, column=col).border = border
        
        # Add classrooms section
        row += 2
        ws_main[f'A{row}'] = "Classroom & Operational Details"
        ws_main[f'A{row}'].font = bold_font
        ws_main[f'A{row}'].fill = header_fill
        row += 1
        
        # Classrooms headers
        headers = ["Classroom Name", "Capacity", "Teacher Ratio", "Avg Students"]
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=row, column=col)
            cell.value = header
            cell.font = bold_font
            cell.fill = subheader_fill
            cell.border = border
        
        row += 1
        
        # Classrooms data
        for classroom in form_data.get("classrooms", []):
            ws_main[f'A{row}'] = classroom.get("name", "")
            ws_main[f'B{row}'] = classroom.get("capacity", 0)
            ws_main[f'C{row}'] = classroom.get("ratio", 0)
            ws_main[f'D{row}'] = classroom.get("avgStudents", 0)
            
            for col in range(1, 5):
                ws_main.cell(row=row, column=col).border = border
            
            row += 1
        
        # Add operating details
        row += 1
        ws_main[f'A{row}'] = "Operating Hours (Daily)"
        ws_main[f'B{row}'] = form_data.get("operatingHours", 0)
        row += 1
        ws_main[f'A{row}'] = "Operating Days (Yearly)"
        ws_main[f'B{row}'] = form_data.get("operatingDays", 0)
        
        # Add business goals
        row += 2
        ws_main[f'A{row}'] = "Business Goals"
        ws_main[f'A{row}'].font = bold_font
        ws_main[f'A{row}'].fill = header_fill
        row += 1
        
        # Goals headers
        headers = ["Goal", "Target Percentage"]
        for col, header in enumerate(headers, 1):
            cell = ws_main.cell(row=row, column=col)
            cell.value = header
            cell.font = bold_font
            cell.fill = subheader_fill
            cell.border = border
        
        row += 1
        
        # Goals data
        for goal in form_data.get("goals", []):
            ws_main[f'A{row}'] = goal.get("goal", "")
            ws_main[f'B{row}'] = goal.get("targetPercentage", 0)
            
            for col in range(1, 3):
                ws_main.cell(row=row, column=col).border = border
            
            row += 1
        
        # Auto-adjust column widths
        for column in ws_main.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 30)
            ws_main.column_dimensions[column_letter].width = adjusted_width
        
        # Save workbook to bytes
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return Excel file
        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": "attachment; filename=daycare_budget_planner.xlsx"}
        )
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating Excel report: {e}")
    


@insight_router.post("/pro-forma-dashboard-data")
async def get_pro_forma_dashboard_data(
                                input_data: dict,
                                token_payload: dict = Depends(verify_token),
                                db: aiomysql.Connection = Depends(AuthDatabaseService.get_db)
                            ):
    try:
        # Token payload contains decoded JWT claims
        async with db.cursor(aiomysql.DictCursor) as cursor:
            # Check if the user exists based on email
            await cursor.execute(
                """
                SELECT * 
                FROM auth_users WHERE id = %s
                """,
                (token_payload.get("sub"),)
            )
        user = await cursor.fetchone()  # Fetch the first matching user
        if not user:
            raise HTTPException(
                status_code=401,
                detail="Invalid token: user ID not found"
            )
        conn = await AuthDatabaseService.connection()
        try:
            async with conn.cursor(aiomysql.DictCursor) as cursor:
                # Convert Pydantic model to dict, then to JSON string
                # Insert user email and input JSON into the table
                await cursor.execute(
                    """
                    SELECT id, user_email, data, created_at
                    FROM input_data 
                    WHERE user_email = %s
                    ORDER BY created_at DESC
                    LIMIT 1
                    """,
                    (user["email"],)
                )
            row = await cursor.fetchone()  # Fetch the first matching user
            if not row:
                raise HTTPException(status_code=404, detail="No inputs found for this user.")
            
            input_data = json.loads(row['data'])
        except Exception as e:
            raise HTTPException(status_code=500, detail=f"Error on input data: {e}")
        finally:
            conn.close()

        # input_data = {"businessName": "Demo Daycare Center", 
        #               "revenueSources": [{"id": "1", "sourceName": "Tution", "monthlyAmount": 50000.0, "tag": ""}, 
        #                                  {"id": "2", "sourceName": "other", "monthlyAmount": 5.0, "tag": ""}], 
        #               "employees": [{"id": "1", "expenseName": "teacher", "monthlyAmount": 5000.0, "type": "Monthly", "hoursPerMonth": 0}, 
        #                             {"id": "2", "expenseName": "director", "monthlyAmount": 10000.0, "type": "Monthly", "hoursPerMonth": 0}], 
        #               "facilities": [{"id": "1", "expenseName": "house", "monthlyAmount": 2000.0, "type": "Monthly"},
        #                              {"id": "1", "expenseName": "other", "monthlyAmount": 200.0, "type": "Monthly"}], 
        #               "administrative": [{"id": "1", "expenseName": "software", "monthlyAmount": 500.0, "type": "Monthly"},
        #                                  {"id": "1", "expenseName": "", "monthlyAmount": 0.0, "type": "Monthly"}], 
        #               "supplies": [{"id": "1", "expenseName": "utilites", "monthlyAmount": 300.0, "type": "Monthly"},
        #                            {"id": "1", "expenseName": "other", "monthlyAmount": 0.0, "type": "Monthly"}], 
        #               "classrooms": [{"id": "1", "name": "todler", "capacity": 23, "ratio": 6.0, "avgStudents": 12},
        #                              {"id": "1", "name": "other", "capacity": 2, "ratio": 5.0, "avgStudents": 1}], 
        #               "operatingHours": 123.0, "operatingDays": 122.0, 
        #               "goals": [{"id": "1", "goal": "Increase Revenue", "targetPercentage": 15.0}, 
        #                         {"id": "2", "goal": "Reduce Expense", "targetPercentage": 20.0}, 
        #                         {"id": "2", "goal": "Improve Classroom Utilization", "targetPercentage": 30.0}]}
        
        # Calculate current year totals
        current_year_revenue = sum(item.get("monthlyAmount", 0) for item in input_data.get("revenueSources", [])) * 12
        current_year_expenses = 0
        
        # Calculate expenses from all categories
        expense_categories = ["employees", "facilities", "administrative", "supplies"]
        for category in expense_categories:
            current_year_expenses += sum(item.get("monthlyAmount", 0) for item in input_data.get(category, [])) * 12
        
        current_year_profit = current_year_revenue - current_year_expenses
        
        # Calculate classroom utilization
        total_capacity = sum(classroom.get("capacity", 0) for classroom in input_data.get("classrooms", []))
        total_students = sum(classroom.get("avgStudents", 0) for classroom in input_data.get("classrooms", []))
        current_utilization = (total_students / total_capacity * 100) if total_capacity > 0 else 0
        
        # Get goal percentages
        goals = input_data.get("goals", [])
        revenue_increase_goal = next((goal for goal in goals if goal.get("goal") == "Increase Revenue"), {})
        expense_reduction_goal = next((goal for goal in goals if goal.get("goal") == "Reduce Expense"), {})
        utilization_goal = next((goal for goal in goals if goal.get("goal") == "Improve Classroom Utilization"), {})
        
        revenue_increase_pct = revenue_increase_goal.get("targetPercentage", 0)
        expense_reduction_pct = expense_reduction_goal.get("targetPercentage", 0)
        utilization_increase_pct = utilization_goal.get("targetPercentage", 0)
        
        # Calculate next year projections
        next_year_revenue = current_year_revenue * (1 + revenue_increase_pct / 100)
        next_year_expenses = current_year_expenses * (1 - expense_reduction_pct / 100)
        next_year_profit = next_year_revenue - next_year_expenses
        next_year_utilization = min(current_utilization * (1 + utilization_increase_pct / 100), 100)
        
        # Generate line chart data (24 months)
        line_chart_data = []
        monthly_revenue_growth = (next_year_revenue - current_year_revenue) / 12
        monthly_expense_reduction = (current_year_expenses - next_year_expenses) / 12
        monthly_utilization_growth = (next_year_utilization - current_utilization) / 12
        
        current_date = datetime.now()
        current_year = current_date.year
        
        for month in range(24):
            # if month < 12:
            #     # Current year - constant values
            #     revenue = current_year_revenue / 12
            #     expenses = current_year_expenses / 12
            #     utilization = current_utilization
            # else:
            #     # Next year - linear progression
            #     progress = (month - 11) / 12  # From 0 to 1 over 12 months
            #     revenue = (current_year_revenue / 12) + (monthly_revenue_growth * progress)
            #     expenses = (current_year_expenses / 12) - (monthly_expense_reduction * progress)
            #     utilization = current_utilization + (monthly_utilization_growth * progress)
            
            # if month < 12:
            #     # 2025: ramp from 0 → current year target
            #     progress = (month + 1) / 12
            #     revenue = current_year_revenue * progress
            #     expenses = current_year_expenses * (1 - (progress * 0.2))  # Example: reduce 20% by Dec
            #     utilization = current_utilization * progress
            # else:
            #     # 2026: smooth transition from 2025 final → 2026 target
            #     progress = (month - 11) / 12
            #     revenue = current_year_revenue + (next_year_revenue - current_year_revenue) * progress
            #     expenses = current_year_expenses - (current_year_expenses - next_year_expenses) * progress
            #     utilization = current_utilization + (next_year_utilization - current_utilization) * progress
            
            if month < 12:
                # 2025: ramp from 0 → current year target
                progress = (month + 1) / 12
                revenue = current_year_revenue * progress
                expenses = current_year_expenses * (1 - (progress * 0.2))  # Example: reduce 20% by Dec
                utilization = current_utilization * progress
            else:
                # 2026: smooth transition from Dec 2025 final → 2026 target
                transition_progress = (month - 11) / 12  # Jan=1/12 ... Dec=12/12

                # Dec 2025 values (end of ramp)
                last_revenue_2025 = current_year_revenue
                last_expenses_2025 = current_year_expenses * 0.8   # since reduced by 20%
                last_utilization_2025 = current_utilization

                # Smoothly interpolate from Dec 2025 → 2026 totals
                revenue = last_revenue_2025 + (next_year_revenue - last_revenue_2025) * transition_progress
                expenses = last_expenses_2025 + (next_year_expenses - last_expenses_2025) * transition_progress
                utilization = last_utilization_2025 + (next_year_utilization - last_utilization_2025) * transition_progress
            
            line_chart_data.append({
                "month": f"{month % 12 + 1}/{current_year + math.floor(month / 12)}",
                "revenue": revenue,
                "expenses": expenses,
                "utilization": utilization
            })
        
        # Generate bar chart data (yearly comparison)
        bar_chart_data = [
            {
                "year": "Current Year",
                "revenue": current_year_revenue,
                "expenses": current_year_expenses,
                "profit": current_year_profit,
                "utilization": current_utilization
            },
            {
                "year": "Next Year",
                "revenue": next_year_revenue,
                "expenses": next_year_expenses,
                "profit": next_year_profit,
                "utilization": next_year_utilization
            }
        ]
        
        # Generate goal-oriented chart data
        goal_chart_data = [
            {
                "goal_type": "Increase Revenue",
                "target_percentage": revenue_increase_pct,
                "achieved_percentage": min(revenue_increase_pct, 100)  # For demo, assume target is achieved
            },
            {
                "goal_type": "Reduce Expense",
                "target_percentage": expense_reduction_pct,
                "achieved_percentage": min(expense_reduction_pct, 100)
            },
            {
                "goal_type": "Improve Utilization",
                "target_percentage": utilization_increase_pct,
                "achieved_percentage": min(utilization_increase_pct, 100)
            }
        ]
        
        return {
            "line_chart": line_chart_data,
            "bar_chart": bar_chart_data,
            "goal_chart": goal_chart_data,
            "summary": {
                "current_year": {
                    "revenue": current_year_revenue,
                    "expenses": current_year_expenses,
                    "profit": current_year_profit,
                    "utilization": current_utilization
                },
                "next_year": {
                    "revenue": next_year_revenue,
                    "expenses": next_year_expenses,
                    "profit": next_year_profit,
                    "utilization": next_year_utilization
                },
                "growth_rates": {
                    "revenue": revenue_increase_pct,
                    "expenses": -expense_reduction_pct,  # Negative because we're reducing expenses
                    "profit": ((next_year_profit - current_year_profit) / current_year_profit * 100) if current_year_profit != 0 else 0,
                    "utilization": utilization_increase_pct
                }
            }
        }
        
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Error generating dashboard data: {e}")
            

from types import SimpleNamespace

def dict_to_obj(d):
    if isinstance(d, dict):
        return SimpleNamespace(**{k: dict_to_obj(v) for k, v in d.items()})
    elif isinstance(d, list):
        return [dict_to_obj(x) for x in d]
    return d

@insight_router.post("/send-email-report")
async def generate_email_report(
                            token_payload: dict = Depends(verify_token),
                            db: aiomysql.Connection = Depends(AuthDatabaseService.get_db)
                        ):
    try:
        # Token payload contains decoded JWT claims
        async with db.cursor(aiomysql.DictCursor) as cursor:
            # Check if the user exists based on email
            await cursor.execute(
                "SELECT * FROM auth_users WHERE id = %s",
                (token_payload.get("sub"),)
            )
        user = await cursor.fetchone()  # Fetch the first matching user
        logger.info(f"User: {user['email']}")
        if not user:
            raise HTTPException(
                status_code=401,
                detail="Invalid token: user ID not found"
            )
        
        async with db.cursor(aiomysql.DictCursor) as cursor:
            # Fetch the latest input_data row for the user
            await cursor.execute(
                """
                SELECT id, user_email, data, created_at 
                FROM insights_data 
                WHERE user_email = %s 
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (user['email'],)
            )
            row = await cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="No insights found for this user.")
            try:
                insight_data = json.loads(row['data'])
            except Exception:
                insight_data = {}
                
        report = dict_to_obj(insight_data)
        
        sender_email = logger_settings.MY_EMAIL
        receiver_email = user['email']
        subject = "Daycare Center Financial Report"

        # Create HTML body
        html_body = f"""
        <html>
        <body>
        <h1 style="color:#1890ff;">Daycare Center Financial Report</h1>

        <h2>Net Monthly Income</h2>
        <p><b>Value:</b> {report.net_monthly_income.value} {report.net_monthly_income.currency}</p>
        <p><b>Calculation:</b> {report.net_monthly_income.calculation}</p>
        <p><b>Note:</b> {report.net_monthly_income.note}</p>

        <h2>Break-Even Enrollment</h2>
        <p><b>Value:</b> {report.break_even_enrollment.value} {report.break_even_enrollment.unit}</p>
        <p><b>Calculation:</b> {report.break_even_enrollment.calculation}</p>
        <p><b>Note:</b> {report.break_even_enrollment.note}</p>

        <h2>Largest Expense</h2>
        <p><b>Category:</b> {report.largest_expense.category}</p>
        <p><b>Percentage of Total Expenses:</b> {report.largest_expense.percentage_of_total_expenses}%</p>
        <p><b>Calculation:</b> {report.largest_expense.calculation}</p>

        <h2>Capacity Utilization</h2>
        <p><b>Value:</b> {report.capacity_utilization.value} {report.capacity_utilization.unit}</p>
        <p><b>Calculation:</b> {report.capacity_utilization.calculation}</p>
        <p><b>Note:</b> {report.capacity_utilization.note}</p>

        <h2>Executive Summary</h2>
        <p><b>Financial Overview:</b> {report.executive_summary.financial_overview}</p>
        <p><b>Profitability Status:</b> {report.executive_summary.profitability_status}</p>
        <p><b>Enrollment Status:</b> {report.executive_summary.enrollment_status}</p>

        <h3>Recommendations:</h3>
        <ul>
        {''.join([f"<li>{rec}</li>" for rec in report.executive_summary.recommendations])}
        </ul>
        </body>
        </html>
        """

        # Create email
        message = MIMEMultipart("alternative")
        message["From"] = sender_email
        message["To"] = receiver_email
        message["Subject"] = subject
        message.attach(MIMEText(html_body, "html"))

        # Send email
        server = smtplib.SMTP("smtp.gmail.com", 587)
        server.starttls()
        server.login(sender_email, logger_settings.EMAIL_APP_PASSWORD)  # replace with your app password
        server.sendmail(sender_email, receiver_email, message.as_string())
        server.quit()

        logger.info(f"Report sent to {receiver_email}")
        return JSONResponse(
                status_code=status.HTTP_200_OK,
                content={"status": "success", "message": f"Report sent to {receiver_email}"}
            )
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            content={"status": "error", "message": f"Failed to send email: {e}"}
        )
        
# -------------------------
# Pydantic Models
# -------------------------
class NetMonthlyIncome(BaseModel):
    value: float
    currency: str
    calculation: str
    note: str

class BreakEvenEnrollment(BaseModel):
    value: int
    unit: str
    calculation: str
    note: str

class LargestExpense(BaseModel):
    category: str
    percentage_of_total_expenses: float
    calculation: str

class CapacityUtilization(BaseModel):
    value: float
    unit: str
    calculation: str
    note: str

class ExecutiveSummary(BaseModel):
    financial_overview: str
    profitability_status: str
    enrollment_status: str
    recommendations: List[str]

class ReportData(BaseModel):
    net_monthly_income: NetMonthlyIncome
    break_even_enrollment: BreakEvenEnrollment
    largest_expense: LargestExpense
    capacity_utilization: CapacityUtilization
    executive_summary: ExecutiveSummary


# -------------------------
# Helper function to create PDF
# -------------------------
async def create_pdf(data: ReportData, filename: str):
    doc = SimpleDocTemplate(filename, pagesize=letter)
    styles = getSampleStyleSheet()
    elements = []

    # Title
    elements.append(Paragraph("Daycare Financial Report", styles["Title"]))
    elements.append(Spacer(1, 20))

    # Section 1
    elements.append(Paragraph("Net Monthly Income", styles["Heading2"]))
    elements.append(Paragraph(f"Value: {data.net_monthly_income.value} {data.net_monthly_income.currency}", styles["Normal"]))
    elements.append(Paragraph(f"Calculation: {data.net_monthly_income.calculation}", styles["Normal"]))
    elements.append(Paragraph(f"Note: {data.net_monthly_income.note}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Section 2
    elements.append(Paragraph("Break-Even Enrollment", styles["Heading2"]))
    elements.append(Paragraph(f"Value: {data.break_even_enrollment.value} {data.break_even_enrollment.unit}", styles["Normal"]))
    elements.append(Paragraph(f"Calculation: {data.break_even_enrollment.calculation}", styles["Normal"]))
    elements.append(Paragraph(f"Note: {data.break_even_enrollment.note}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Section 3
    elements.append(Paragraph("Largest Expense", styles["Heading2"]))
    elements.append(Paragraph(f"Category: {data.largest_expense.category}", styles["Normal"]))
    elements.append(Paragraph(f"Percentage of Total Expenses: {data.largest_expense.percentage_of_total_expenses}%", styles["Normal"]))
    elements.append(Paragraph(f"Calculation: {data.largest_expense.calculation}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Section 4
    elements.append(Paragraph("Capacity Utilization", styles["Heading2"]))
    elements.append(Paragraph(f"Value: {data.capacity_utilization.value} {data.capacity_utilization.unit}", styles["Normal"]))
    elements.append(Paragraph(f"Calculation: {data.capacity_utilization.calculation}", styles["Normal"]))
    elements.append(Paragraph(f"Note: {data.capacity_utilization.note}", styles["Normal"]))
    elements.append(Spacer(1, 12))

    # Section 5
    elements.append(Paragraph("Executive Summary", styles["Heading2"]))
    elements.append(Paragraph(f"Financial Overview: {data.executive_summary.financial_overview}", styles["Normal"]))
    elements.append(Paragraph(f"Profitability Status: {data.executive_summary.profitability_status}", styles["Normal"]))
    elements.append(Paragraph(f"Enrollment Status: {data.executive_summary.enrollment_status}", styles["Normal"]))
    
    elements.append(Paragraph("Recommendations:", styles["Normal"]))
    for rec in data.executive_summary.recommendations:
        elements.append(Paragraph(f"- {rec}", styles["Normal"]))

    # Build PDF
    doc.build(elements)


# -------------------------
# FastAPI Endpoint
# -------------------------
@insight_router.get("/generate-pdf-report")
async def generate_pdf_report(
                token_payload: dict = Depends(verify_token),
                db: aiomysql.Connection = Depends(AuthDatabaseService.get_db)
            ):
    try:
        async with db.cursor(aiomysql.DictCursor) as cursor:
            # Check if the user exists based on email
            await cursor.execute(
                "SELECT * FROM auth_users WHERE id = %s",
                (token_payload.get("sub"),)
            )
        user = await cursor.fetchone()  # Fetch the first matching user
        logger.debug(f"User: {user['email']}")
        if not user:
            raise HTTPException(
                status_code=401,
                detail="Invalid token: user ID not found"
            )
        
        async with db.cursor(aiomysql.DictCursor) as cursor:
            # Fetch the latest input_data row for the user
            await cursor.execute(
                """
                SELECT id, user_email, data, created_at 
                FROM insights_data 
                WHERE user_email = %s 
                ORDER BY created_at DESC
                LIMIT 1
                """,
                (user['email'],)
            )
            row = await cursor.fetchone()
            if not row:
                raise HTTPException(status_code=404, detail="No insights found for this user.")
            try:
                insights = json.loads(row['data'])
            except Exception:
                insights = {}
                
        report_data = dict_to_obj(insights)
        filename = f"report_{uuid.uuid4().hex}.pdf"
        await create_pdf(report_data, filename)
        return FileResponse(
            filename, 
            media_type="application/pdf", 
            filename="daycare_report.pdf"
            )

    except FileNotFoundError as e:
        raise HTTPException(status_code=404, detail=str(e))
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"Unexpected error when generating PDF: {str(e)}")
      

@insight_router.get("/user")
async def get_insights(
                    token_payload: dict = Depends(verify_token),
                    db: aiomysql.Connection = Depends(AuthDatabaseService.get_db)
                    ):
    try:
        async with db.cursor(aiomysql.DictCursor) as cursor:
            # Check if the user exists based on email
            await cursor.execute(
                "SELECT * FROM auth_users WHERE id = %s",
                (token_payload.get("sub"),)
            )
        user = await cursor.fetchone()  # Fetch the first matching user
        logger.debug(f"User: {user['email']}")
        if not user:
            raise HTTPException(
                status_code=401,
                detail="Invalid token: user ID not found"
            )
        return {
                "email": user['email'], 
                "username": user['username']
                }
    except Exception as e:
        return JSONResponse(
            status_code=status.HTTP_511_NETWORK_AUTHENTICATION_REQUIRED,
            content={"status": "error", "message": f"Failed to know user: {e}"}
        )
